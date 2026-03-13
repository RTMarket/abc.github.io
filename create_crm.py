#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建工作簿
wb = Workbook()

# ===== 首页/Cover =====
ws_cover = wb.active
ws_cover.title = "首页"

# 样式定义
header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
header_font = Font(size=14, bold=True, color="FFFFFF")
title_font = Font(size=20, bold=True, color="FF6B00")
subtitle_font = Font(size=12, color="666666")
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 封面内容
ws_cover['B2'] = "香港佳和供应链管理有限公司"
ws_cover['B2'].font = Font(size=24, bold=True, color="FF6B00")
ws_cover['B3'] = "CRM 客户管理系统"
ws_cover['B3'].font = Font(size=18, bold=True, color="333333")
ws_cover['B5'] = "数据汇总表"
ws_cover['B5'].font = title_font

ws_cover['B7'] = "表格说明："
ws_cover['B7'].font = Font(size=12, bold=True)
ws_cover['B8'] = "• 卖家CRM：跨境卖家入驻申请表数据"
ws_cover['B9'] = "• 工厂CRM：海外工厂注册申请表数据"
ws_cover['B10'] = "• 数据汇总：关键指标统计"

ws_cover['B12'] = "使用方法："
ws_cover['B12'].font = Font(size=12, bold=True)
ws_cover['B13'] = "1. 在「卖家CRM」和「工厂CRM」表单中查看客户信息"
ws_cover['B14'] = "2. 可按需筛选、排序和分析数据"
ws_cover['B15'] = "3. 「数据汇总」工作表自动统计关键指标"

ws_cover.column_dimensions['B'].width = 50

# ===== 卖家CRM工作表 =====
ws_seller = wb.create_sheet("卖家CRM")

# 卖家数据表头
seller_headers = [
    "序号", "公司名称(中文)", "公司名称(英文)", "统一社会信用代码",
    "联系人", "电话", "邮箱", "微信",
    "当前平台", "目标平台", "美国店铺数", "墨西哥店铺数",
    "需要合伙人", "需要美国银行", "需要墨西哥银行", "月销售额",
    "产品品类", "海外仓状态", "备货资金",
    "需要供应链", "需要MCN", "需要物流", "需要支付",
    "服务方案", "预算", "其他需求", "提交时间"
]

for col, header in enumerate(seller_headers, 1):
    cell = ws_seller.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin

# 设置列宽
ws_seller.column_dimensions['A'].width = 8
ws_seller.column_dimensions['B'].width = 25
ws_seller.column_dimensions['C'].width = 25
ws_seller.column_dimensions['D'].width = 18
ws_seller.column_dimensions['E'].width = 15
ws_seller.column_dimensions['F'].width = 15
ws_seller.column_dimensions['G'].width = 25
ws_seller.column_dimensions['H'].width = 20
ws_seller.column_dimensions['I'].width = 30
ws_seller.column_dimensions['J'].width = 30
ws_seller.column_dimensions['K'].width = 12
ws_seller.column_dimensions['L'].width = 12
ws_seller.column_dimensions['M'].width = 12
ws_seller.column_dimensions['N'].width = 12
ws_seller.column_dimensions['O'].width = 12
ws_seller.column_dimensions['P'].width = 15
ws_seller.column_dimensions['Q'].width = 35
ws_seller.column_dimensions['R'].width = 15
ws_seller.column_dimensions['S'].width = 15
ws_seller.column_dimensions['T'].width = 12
ws_seller.column_dimensions['U'].width = 12
ws_seller.column_dimensions['V'].width = 12
ws_seller.column_dimensions['W'].width = 12
ws_seller.column_dimensions['X'].width = 15
ws_seller.column_dimensions['Y'].width = 15
ws_seller.column_dimensions['Z'].width = 30
ws_seller.column_dimensions['AA'].width = 18

# 示例数据行（2-11行）- 模拟数据
sample_seller_data = [
    [1, "深圳跨境电商有限公司", "Shenzhen Cross-Border E-commerce Co., Ltd.", "91440300MA5EXAMPLE",
     "张先生", "+86 13800138000", "zhang@company.com", "Zhang138001",
     "Amazon,TikTok Shop", "TEMU,SHEIN", "1-3", "0",
     "是", "是", "否", "10-30万",
     "汽摩配件,家居智能", "计划备货", "10-50万",
     "需要", "不需要", "需要", "需要",
     "方案B（进阶版）", "5-10万", "需要美国本土合规指导", "2024-01-15 10:30"],
    [2, "浙江供应链管理企业", "Zhejiang Supply Chain Management Co., Ltd.", "91330000MA5EXAMPLE",
     "李女士", "+86 13900139000", "li@company.com", "Li139001",
     "TEMU", "Amazon,TikTok Shop", "0", "1-3",
     "是", "是", "是", "30万以上",
     "母婴用品,儿童玩具", "已有海外仓", "50万以上",
     "需要", "需要", "需要", "需要",
     "方案C（终极版）", "10-30万", "希望尽快开拓北美市场", "2024-01-16 14:20"],
    [3, "广州贸易公司", "Guangzhou Trading Company", "91440000MA5EXAMPLE",
     "王先生", "+86 13700137000", "wang@company.com", "Wang137001",
     "SHEIN", "TikTok Shop,Amazon", "4-10", "1-3",
     "否", "是", "否", "5-10万",
     "日用家居,五金工具", "计划备货", "1-10万",
     "需要", "不需要", "需要", "否",
     "方案A（基础版）", "1-5万", "", "2024-01-17 09:15"],
]

for row_idx, row_data in enumerate(sample_seller_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_seller.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ===== 工厂CRM工作表 =====
ws_factory = wb.create_sheet("工厂CRM")

# 工厂数据表头
factory_headers = [
    "序号", "工厂名称(英文)", "工厂名称(中文)", "注册国家", "注册地址", "税号",
    "联系人", "电话", "邮箱", "网址",
    "产品品类", "产品描述", "最小订单量", "交货周期", "产能",
    "是否有认证", "认证证书", "本地库存", "物流方式",
    "接受样品单", "品牌服务", "其他信息", "提交时间"
]

for col, header in enumerate(factory_headers, 1):
    cell = ws_factory.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin

# 设置列宽
ws_factory.column_dimensions['A'].width = 8
ws_factory.column_dimensions['B'].width = 30
ws_factory.column_dimensions['C'].width = 25
ws_factory.column_dimensions['D'].width = 12
ws_factory.column_dimensions['E'].width = 35
ws_factory.column_dimensions['F'].width = 20
ws_factory.column_dimensions['G'].width = 15
ws_factory.column_dimensions['H'].width = 15
ws_factory.column_dimensions['I'].width = 25
ws_factory.column_dimensions['J'].width = 25
ws_factory.column_dimensions['K'].width = 35
ws_factory.column_dimensions['L'].width = 30
ws_factory.column_dimensions['M'].width = 12
ws_factory.column_dimensions['N'].width = 12
ws_factory.column_dimensions['O'].width = 12
ws_factory.column_dimensions['P'].width = 12
ws_factory.column_dimensions['Q'].width = 30
ws_factory.column_dimensions['R'].width = 18
ws_factory.column_dimensions['S'].width = 25
ws_factory.column_dimensions['T'].width = 12
ws_factory.column_dimensions['U'].width = 12
ws_factory.column_dimensions['V'].width = 35
ws_factory.column_dimensions['W'].width = 18

# 示例数据 - 模拟数据
sample_factory_data = [
    [1, "USA Manufacturing Co.", "美国制造有限公司", "USA", "123 Factory Ave, Los Angeles, CA", "12-3456789",
     "John Smith", "+1 555-123-4567", "john@usafactory.com", "www.usafactory.com",
     "汽摩配件,五金工具", "专业生产汽车零部件和工具", "500件", "15-20天", "月产50000件",
     "是", "FCC,UL,ISO9001", "美国本土仓", "一件代发,FBA补货",
     "接受", "有", "专注北美市场多年，与多家电商合作", "2024-01-10 11:00"],
    [2, "Mexico Industrial S.A. de C.V.", "墨西哥工业有限公司", "Mexico", "Av. Industrial 500, Mexico City", "MIX123456789",
     "Carlos Garcia", "+52 55 1234 5678", "carlos@mxindustrial.com", "",
     "医疗保健,实验室耗材", "医疗器械和实验室设备生产", "1000件", "20-30天", "月产20000件",
     "是", "CE,FDA,ISO13485", "墨西哥本土仓", "本土仓发货,大宗贸易",
     "接受", "无", "工厂位于墨西哥城工业区，交通便利", "2024-01-12 16:30"],
    [3, "Smart Home Tech Ltd.", "智能家居科技有限公司", "Other", "Shenzhen, Guangdong, China", "91440300MA5EXAMPLE",
     "张伟", "+86 13600136000", "zhangwei@smarthome.com", "www.smarthome-tech.com",
     "家居智能,宠物智能用品", "智能家居产品研发生产", "200件", "7-10天", "月产100000件",
     "是", "CE,RoHS,FCC", "暂无本土库存", "一件代发",
     "接受", "有", "可提供OEM/ODM服务", "2024-01-18 13:45"],
]

for row_idx, row_data in enumerate(sample_factory_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_factory.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ===== 数据汇总工作表 =====
ws_summary = wb.create_sheet("数据汇总")

# 汇总标题
ws_summary['B2'] = "数据统计汇总"
ws_summary['B2'].font = Font(size=18, bold=True, color="FF6B00")
ws_summary['B3'] = f"统计时间: {pd.Timestamp.now().strftime('%Y-%m-%d')}"
ws_summary['B3'].font = subtitle_font

# 卖家统计
ws_summary['B5'] = "【卖家数据统计】"
ws_summary['B5'].font = Font(size=14, bold=True)
ws_summary['B5'].fill = PatternFill(start_color="FFE4D6", end_color="FFE4D6", fill_type="solid")

seller_stats = [
    ["指标", "数值"],
    ["总卖家数", "=COUNTA(卖家CRM!A2:A1000)"],
    ["有美国店铺的卖家", "=COUNTIF(卖家CRM!K2:K1000,\"<>0\")"],
    ["有墨西哥店铺的卖家", "=COUNTIF(卖家CRM!L2:L1000,\"<>0\")"],
    ["需要合伙人服务", "=COUNTIF(卖家CRM!M2:M1000,\"是\")"],
    ["需要美国银行账户", "=COUNTIF(卖家CRM!N2:N1000,\"是\")"],
    ["需要墨西哥银行账户", "=COUNTIF(卖家CRM!O2:O1000,\"是\")"],
    ["需要供应链服务", "=COUNTIF(卖家CRM!T2:T1000,\"需要\")"],
    ["需要物流服务", "=COUNTIF(卖家CRM!V2:V1000,\"需要\")"],
    ["需要支付服务", "=COUNTIF(卖家CRM!W2:W1000,\"需要\")"],
]

for row_idx, (label, formula) in enumerate(seller_stats, 7):
    ws_summary.cell(row=row_idx, column=2).value = label
    ws_summary.cell(row=row_idx, column=3).value = formula
    ws_summary.cell(row=row_idx, column=2).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws_summary.cell(row=row_idx, column=2).border = border_thin
    ws_summary.cell(row=row_idx, column=3).border = border_thin

# 工厂统计
ws_summary['B19'] = "【工厂数据统计】"
ws_summary['B19'].font = Font(size=14, bold=True)
ws_summary['B19'].fill = PatternFill(start_color="FFE4D6", end_color="FFE4D6", fill_type="solid")

factory_stats = [
    ["指标", "数值"],
    ["总工厂数", "=COUNTA(工厂CRM!A2:A1000)"],
    ["有认证的工厂", "=COUNTIF(工厂CRM!P2:P1000,\"是\")"],
    ["有美国本土仓", "=COUNTIF(工厂CRM!R2:R1000,\"*美国*\")"],
    ["有墨西哥本土仓", "=COUNTIF(工厂CRM!R2:R1000,\"*墨西哥*\")"],
    ["接受样品单", "=COUNTIF(工厂CRM!T2:T1000,\"接受\")"],
    ["提供品牌服务", "=COUNTIF(工厂CRM!U2:U1000,\"有\")"],
]

for row_idx, (label, formula) in enumerate(factory_stats, 21):
    ws_summary.cell(row=row_idx, column=2).value = label
    ws_summary.cell(row=row_idx, column=3).value = formula
    ws_summary.cell(row=row_idx, column=2).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws_summary.cell(row=row_idx, column=2).border = border_thin
    ws_summary.cell(row=row_idx, column=3).border = border_thin

ws_summary.column_dimensions['B'].width = 25
ws_summary.column_dimensions['C'].width = 15

# 隐藏网格线
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

# 保存文件
output_path = "/workspace/company-portal/CRM客户管理系统.xlsx"
wb.save(output_path)
print(f"CRM表格已创建: {output_path}")
